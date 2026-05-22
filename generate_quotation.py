import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Add skill scripts to path
sys.path.append(r'C:\Users\Admin\.accio\accounts\1751070282\skills\xlsx\scripts')
from builders import build_table
from style_kit import NUMFMT

def generate():
    wb = Workbook()
    # Use a CJK-capable font as recommended
    wb._named_styles['Normal'].font = Font(name='Microsoft YaHei', size=11)
    ws = wb.active
    ws.title = 'Quotation'

    # 1. Header Information
    ws.merge_cells('A1:E1')
    ws['A1'] = 'QUOTATION'
    ws['A1'].font = Font(size=20, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Seller Info
    ws['A3'] = 'Seller:'
    ws['B3'] = 'Shandong Forgealloy Racing Tech Co., Ltd'
    ws['A4'] = 'Date:'
    ws['B4'] = '2026-05-22'
    ws['A5'] = 'Quotation No:'
    ws['B5'] = 'Q20260522-01'

    # Bill To
    ws['A7'] = 'Bill To:'
    ws['B7'] = 'MOZA CONSULTORIA LTDA'
    ws['B8'] = 'Mozaniel Andrade de Souza'
    ws['B9'] = 'rua Raposo da Fonseca, 1014, São Paulo, SP, Brasil'
    ws['B10'] = '+5511984387707'

    # 2. Product Table
    # A12 is header row
    data = [
        ['Item', 'Description', 'Qty', 'Unit Price ($)', 'Total ($)'],
        [1, 'Customized Gift Cup Set (Boxed)', 100, 2.493, '=C13*D13'],
        [2, 'Customized Ceramic Cup and Saucer (300ml)', 50, 2.99, '=C14*D14'],
        [3, 'Customized Ceramic Cup and Saucer (70ml)', 50, 1.94, '=C15*D15'],
        [4, 'Customized Ceramic Cup and Saucer (200ml)', 50, 2.39, '=C16*D16'],
    ]

    t = build_table(
        ws,
        data=data,
        anchor='A12',
        theme='modern_finance',
        column_formats={
            'Qty': 'int',
            'Unit Price ($)': 'dollar_2',
            'Total ($)': 'dollar_2'
        },
        total_row=True,
        auto_width=True
    )

    # 3. Footer / Terms
    last_row = t['total_row'] + 2
    ws[f'A{last_row}'] = 'Trade Terms:'
    ws[f'B{last_row}'] = 'EXW'
    ws[f'A{last_row+1}'] = 'Validity:'
    ws[f'B{last_row+1}'] = '30 Days'

    file_name = 'Quotation_MOZA_CONSULTORIA_20260522.xlsx'
    wb.save(file_name)
    print(f"File saved as {file_name}")

if __name__ == "__main__":
    generate()
